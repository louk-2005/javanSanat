# django files
from django.contrib import admin
from django.utils.html import format_html

#your files
from .models import  Location, CommunicationWithUs

@admin.register(CommunicationWithUs)
class CommunicationWithUsAdmin(admin.ModelAdmin):
    # فیلدهایی که در لیست نمایش داده می‌شوند
    list_display = ('full_name', 'email', 'phone', 'message_preview', 'is_read_status', 'created_at')

    # فیلدهای قابل جستجو
    search_fields = ('full_name', 'email', 'phone', 'message')

    # فیلترهای سمت راست صفحه
    list_filter = ('created_at', 'phone', 'is_read')

    # مرتب‌سازی پیش‌فرض
    ordering = ('-created_at',)

    # تعداد آیتم‌ها در هر صفحه
    list_per_page = 25

    # فیلدهای فقط خواندنی
    readonly_fields = ('created_at',)

    # سفارشی‌سازی ستون‌ها
    def message_preview(self, obj):
        """نمایش پیش‌نمایش پیام با محدودیت کاراکتر"""
        if len(obj.message) > 50:
            return f"{obj.message[:50]}..."
        return obj.message

    message_preview.short_description = 'پیام'

    def created_at(self, obj):
        """نمایش تاریخ ایجاد به صورت فارسی"""
        if hasattr(obj, 'created_at') and obj.created_at:
            return obj.created_at.strftime("%Y/%m/%d %H:%M")
        return "-"

    created_at.short_description = 'تاریخ ثبت'

    def phone(self, obj):
        """قالب‌بندی شماره تلفن"""
        if obj.phone:
            return format_html(
                '<span style="direction: ltr; display: inline-block;">{}</span>',
                obj.phone
            )
        return "-"

    phone.short_description = 'تلفن'

    def is_read_status(self, obj):
        """نمایش وضعیت خوانده شده با آیکون"""
        if obj.is_read:
            return format_html(
                '<span style="color: green;"><i class="fas fa-check-circle"></i> خوانده شده</span>'
            )
        return format_html(
            '<span style="color: red;"><i class="fas fa-times-circle"></i> خوانده نشده</span>'
        )

    is_read_status.short_description = 'وضعیت'
    is_read_status.admin_order_field = 'is_read'

    # سفارشی‌سازی هدر ادمین
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['title'] = "مدیریت پیام‌های کاربران"
        return super().changelist_view(request, extra_context=extra_context)

    # سفارشی‌سازی فرم ویرایش
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['full_name'].label = "نام کامل"
        form.base_fields['email'].label = "ایمیل"
        form.base_fields['phone'].label = "شماره تلفن"

        # بررسی وجود فیلد message قبل از تغییر برچسب
        if 'message' in form.base_fields:
            form.base_fields['message'].label = "متن پیام"

        # تغییر برچسب فیلد is_read
        if 'is_read' in form.base_fields:
            form.base_fields['is_read'].label = "وضعیت خوانده شده"

        return form

    # اضافه کردن دکمه‌های عملیات انبوه
    actions = ['mark_as_read', 'mark_as_unread', 'export_to_excel']

    def mark_as_read(self, request, queryset):
        """علامت‌گذاری پیام‌ها به عنوان خوانده شده"""
        updated = queryset.update(is_read=True)
        self.message_user(
            request,
            f"{updated} پیام به عنوان خوانده شده علامت‌گذاری شدند.",
            'success'
        )

    mark_as_read.short_description = "علامت‌گذاری به عنوان خوانده شده"

    def mark_as_unread(self, request, queryset):
        """علامت‌گذاری پیام‌ها به عنوان خوانده نشده"""
        updated = queryset.update(is_read=False)
        self.message_user(
            request,
            f"{updated} پیام به عنوان خوانده نشده علامت‌گذاری شدند.",
            'success'
        )

    mark_as_unread.short_description = "علامت‌گذاری به عنوان خوانده نشده"

    def export_to_excel(self, request, queryset):
        """خروجی Excel از پیام‌های انتخاب شده"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
        except ImportError:
            self.message_user(request, "برای خروجی اکسل باید کتابخانه openpyxl نصب شود: pip install openpyxl", 'error')
            return

        # ایجاد یک Workbook جدید
        wb = Workbook()
        ws = wb.active
        ws.title = "پیام‌های کاربران"

        # تعریف استایل‌ها
        header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # تعریف حاشیه‌ها
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # نوشتن هدر جدول
        headers = ['نام کامل', 'ایمیل', 'شماره تلفن', 'متن پیام', 'وضعیت خوانده شده', 'تاریخ ثبت']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # نوشتن داده‌ها
        for row_num, message in enumerate(queryset, 2):
            # نام کامل
            cell = ws.cell(row=row_num, column=1, value=message.full_name)
            cell.border = thin_border

            # ایمیل
            cell = ws.cell(row=row_num, column=2, value=message.email)
            cell.border = thin_border

            # شماره تلفن
            cell = ws.cell(row=row_num, column=3, value=message.phone or "")
            cell.border = thin_border

            # متن پیام
            cell = ws.cell(row=row_num, column=4, value=message.message)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # وضعیت خوانده شده
            status = "خوانده شده" if message.is_read else "خوانده نشده"
            cell = ws.cell(row=row_num, column=5, value=status)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # تاریخ ثبت
            date_value = message.created_at.strftime("%Y/%m/%d %H:%M") if message.created_at else ""
            cell = ws.cell(row=row_num, column=6, value=date_value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # تنظیم عرض ستون‌ها
        column_widths = [20, 25, 15, 50, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # تنظیم ارتفاع ردیف هدر
        ws.row_dimensions[1].height = 25

        # ایجاد پاسخ HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="messages.xlsx"'

        # ذخیره Workbook در پاسخ
        wb.save(response)

        return response

    export_to_excel.short_description = "خروجی Excel از پیام‌ها"


    # سفارشی‌سازی نمایش جزئیات
    fieldsets = (
        ('اطلاعات کاربر', {
            'fields': ('full_name', 'email', 'phone')
        }),
        ('محتوای پیام', {
            'fields': ('message',),
            'classes': ('wide',),
            'description': 'متن کامل پیام ارسالی توسط کاربر'
        }),
        ('وضعیت پیام', {
            'fields': ('is_read',),
            'classes': ('collapse',),
            'description': 'وضعیت خوانده شدن پیام توسط ادمین'
        }),
        ('اطلاعات سیستمی', {
            'fields': ('created_at',),
            'classes': ('collapse',),
            'description': 'اطلاعات ثبت پیام در سیستم'
        }),
    )

    # اضافه کردن کلاس‌های CSS برای بهبود ظاهر
    class Media:
        css = {
            'all': ('https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.4/css/all.min.css',)
        }


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    # List view configuration
    list_display = ('name', 'image_thumbnail', 'coordinates', 'short_description')
    list_filter = ('name',)
    search_fields = ('name', 'description')
    ordering = ('name',)

    # Detail view configuration
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'description')
        }),
        ('Location Data', {
            'fields': ('latitude', 'longitude'),
            'classes': ('collapse',)
        }),
        ('Image', {
            'fields': ('image',),
            'description': 'Image will be automatically resized to 220x120px on save'
        }),
    )

    readonly_fields = ('image_preview',)

    # Custom methods for display
    def image_thumbnail(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" width="50" height="50" style="object-fit: cover; border-radius: 4px;" />',
                obj.image.url
            )
        return "No Image"

    image_thumbnail.short_description = 'Image'

    def coordinates(self, obj):
        return f"{obj.latitude}, {obj.longitude}"

    coordinates.short_description = 'Coordinates'
    coordinates.admin_order_field = 'latitude'

    def short_description(self, obj):
        return (obj.description[:75] + '...') if len(obj.description) > 75 else obj.description

    short_description.short_description = 'Description'

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" width="300" style="max-height: 200px; object-fit: contain; border: 1px solid #eee; border-radius: 4px;" />',
                obj.image.url
            )
        return "Upload an image to see preview"

    image_preview.short_description = 'Current Image'

    # Form customization
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['latitude'].widget.attrs['placeholder'] = 'e.g., 35.689487'
        form.base_fields['longitude'].widget.attrs['placeholder'] = 'e.g., 51.389172'
        return form

    # Removed date_hierarchy since it requires a DateField/DateTimeField
    save_on_top = True